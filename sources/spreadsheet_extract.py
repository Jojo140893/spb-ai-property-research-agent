"""
Adaptive spreadsheet/PDF stocklist extractor.

E-Agent distributes stock as downloadable files rather than HTML listings, and
every builder formats theirs differently. Observed live on e-agent.com.au:

  * "VIC Regional"  — one column of fixed-width text
                      ('Lot 82 Aberdeen      282   142.8  12.0  Sep-26 ...')
  * "NSW Dual"      — a proper multi-column grid with Status/AVAILABLE rows
  * "DUAL QLD"      — multi-column, builder + region + estate header rows

Rather than hand-mapping each, every row is flattened to one text line and fed to
the SAME field parser used for HTML listings (adaptive_extract.parse_fields), so
prices, sizes, bed/bath/car, suburb and titles are recognised wherever they sit.
Group/estate header rows (no price, few numbers) are remembered and used as
context for the lot rows beneath them.

Per-row hyperlinks are kept. A stocklist's "Download" cell carries the link to
that lot's own flyer or floorplan, which is the only per-listing link these files
contain — the file URL itself is shared by every row in the file. In XLSX that
link lives on `cell.hyperlink` (or inside a `=HYPERLINK()` formula); in PDF it is
a page annotation, matched to a row by vertical position.

Nothing is invented: a row without a recognisable price is skipped, and fields
that cannot be found stay None.
"""

import csv
import io
import logging
import re
from typing import Any, Dict, Iterator, List, Optional, Tuple

from sources.adaptive_extract import parse_fields, _clean
from sources.feature_extract import parse_listing_features

logger = logging.getLogger("spb.extract.sheet")

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:  # pragma: no cover
    XLSX_AVAILABLE = False

try:
    import pdfplumber
    PDF_AVAILABLE = True
except ImportError:  # pragma: no cover
    PDF_AVAILABLE = False

LOT_TOKEN = re.compile(r"\blot\s*\d+", re.I)
# Estate/region/builder header lines look like 'ʊ  Aberdeen - Winter Valley - VIC - House & Land'
HEADER_JUNK = re.compile(r"^[^\w]*", re.M)

# A link is only labelled a floorplan/brochure when it says so. Anything else is
# the lot's own page or file, which is still far more use than the shared
# stocklist URL currently stored on every row.
_LINK_KINDS = (
    ("floorplan_url", re.compile(r"floor\s*-?\s*plan|floorplan|site\s*plan", re.I)),
    ("brochure_url", re.compile(r"brochure|flyer|booklet|fact\s*sheet|spec\w*sheet", re.I)),
)
# openpyxl loses a =HYPERLINK() target when the workbook is opened for values.
_HYPERLINK_FN = re.compile(r'HYPERLINK\(\s*"([^"]+)"', re.I)
LinkList = List[Tuple[str, str]]          # [(anchor text, url), ...]


def _row_to_text(cells: List[Any]) -> str:
    """Flatten a spreadsheet row to a single normalised text line."""
    parts = []
    for v in cells:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return _clean(" ".join(parts))


def _classify_link(url: str, anchor: str = "") -> str:
    for key, rx in _LINK_KINDS:
        if rx.search(f"{anchor} {url}"):
            return key
    return "listing_url"


def _link_fields(links: LinkList) -> Dict[str, str]:
    """First link of each kind wins — stocklists list them left to right."""
    out: Dict[str, str] = {}
    for anchor, url in links:
        if url:
            out.setdefault(_classify_link(url, anchor), url)
    return out


STREET_RE = re.compile(
    r"\b\d+[A-Za-z]?(?:[/-]\d+[A-Za-z]?)?\s+[A-Z][A-Za-z'\-]+(?:\s+[A-Z][A-Za-z'\-]+)?\s+"
    r"(?:Street|St|Road|Rd|Avenue|Ave|Drive|Dr|Court|Ct|Place|Pl|Crescent|Cres|Way|"
    r"Circuit|Cct|Parade|Pde|Boulevard|Blvd|Terrace|Tce|Close|Cl|Lane|Ln|Grove|Gr|"
    r"Rise|Loop|Esplanade|Esp|Walk|Mews|Green)\b\.?")


# Several per-builder stocklists open with a bare lot number rather than "Lot N"
# — Aldrich "103 Samara Estate Fraser Rise ...", Hermitage "1123 COMPLETED HOME 444
# ...". Only read it as a lot when the following word is not a street type (that case
# is a street number, and STREET_RE has already claimed it) and not a status or month.
_LEADING_LOT = re.compile(r"^(\d{1,5})[A-Za-z]?\s+(?=[A-Za-z])")
_NOT_AN_ESTATE_WORD = re.compile(
    r"^(available|not|unavailable|under|on|hold|sold|reserved|titled|untitled|registered|"
    r"completed|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|single|double|house|land|"
    r"eoi|new|lot|stage|from|price|total)$", re.I)


def _leading_lot(text: str) -> Optional[str]:
    m = _LEADING_LOT.match(_clean(text))
    if not m or STREET_RE.search(_clean(text)[:60]):
        return None
    nxt = _clean(text)[m.end():].split(" ")[0]
    return m.group(1) if nxt and not _NOT_AN_ESTATE_WORD.match(nxt) else None


_PRICE_ANY = re.compile(r"\$\s?[\d,\.]+")
_COMPACT_OK = re.compile(r"^(?=.*[A-Za-z])[A-Za-z0-9'\-/&,. ]{4,55}$")
_PLACE_SEG = re.compile(r"^[A-Za-z][A-Za-z'\-]*(?: [A-Za-z'\-]+){0,2}(?: \d{4})?$")
# Gallery Group's sheet puts its own title in the row: "100 Park Royal Crescent,
# GALLERY STOCK LIST 2026". That is not a locality.
_NOT_A_PLACE = re.compile(r"\b(stock|list|price\w*|package\w*|availabilit\w*|"
                          r"release|stage|status|update[d]?|sheet|report)\b", re.I)
# Words that mark a stocklist ROW rather than an address. Without these, a jammed row
# that happens to be short once its prices are removed ("103 Samara Estate Fraser Rise
# Available Bristol 15") passes for an address and lands in the client's sheet whole.
_ROW_NOT_ADDRESS = re.compile(
    r"\b(available|unavailable|under\s*offer|on\s*hold|hold|sold|reserved|leased|tbc|"
    r"titled|untitled|registered|single|double|storey|completed|eoi|"
    r"house\s*&?\s*land|dual\s*key)\b", re.I)


def _already_an_address(text: str) -> Optional[str]:
    """Some builders' PDFs give a row that IS just an address ("Lot 444, Hillcrest",
    "4 Windsor Street, BRISBANE NORTH"). Composing a label from parts would only throw
    half of it away, so keep it — minus the price."""
    s = re.sub(r"\s+,", ",", _clean(_PRICE_ANY.sub(" ", text))).strip(" ,-")
    if len(s) > 55 or not _COMPACT_OK.match(s) or _ROW_NOT_ADDRESS.search(s):
        return None
    # must actually look addressed: a leading number, a "Lot N", or a street type.
    # Otherwise marketing copy ("Package from $650,000 enquire now") qualifies.
    if not (re.match(r"^\d", s) or re.search(r"\blot\s*\d", s, re.I) or STREET_RE.search(s)):
        return None
    # a lot or street number, plus maybe a postcode, is fine; a grid of figures is not
    return s if len(re.findall(r"(?<![\w/])\d+(?:\.\d+)?(?!\w)", s)) <= 2 else None


def _place_suffix(text: str, after: int) -> Optional[str]:
    """The first place-looking comma segment after a street match, so "520 Stony Drive"
    keeps its ", Alluvium" instead of being trimmed to the street alone."""
    for seg in _PRICE_ANY.sub(" ", text[after:]).split(","):
        seg = _clean(seg).strip(" -")
        if seg and _PLACE_SEG.match(seg) and not _NOT_A_PLACE.search(seg):
            return seg
    return None


def _address_label(text: str, fields: dict, context: str = "") -> Optional[str]:
    """A short, human-readable label for the client's sheet, or None if the row
    offers nothing better than its own raw text.

    Coleen's complaint was that the address column held the entire stocklist row —
    `parse_fields` sets `lot_address` to the whole line containing "Lot N", which for
    a flattened spreadsheet row is everything. Prefer "105 Almond Street, Denman" or
    "Lot 82, Aberdeen". The full row is never lost: it is kept in `source_text`.
    """
    parts = []
    street = STREET_RE.search(text)
    lot = fields.get("lot_number") or _leading_lot(text)
    if street:
        parts.append(_clean(street.group(0)))
        place = _place_suffix(text, street.end())
        if place:
            parts.append(place)
            return ", ".join(parts)
    elif (compact := _already_an_address(text)):
        return compact
    elif lot:
        parts.append(f"Lot {lot}" if not str(lot).lower().startswith(("lot", "cc-")) else str(lot))
    estate = _clean(re.sub(r"^[^A-Za-z0-9]+", "",
                           str(fields.get("estate_name") or _estate_from_context(context) or "")))
    if parts and estate and 2 < len(estate) <= 44 and "$" not in estate \
            and estate.lower() not in parts[0].lower():
        parts.append(estate)
    return ", ".join(parts) if parts else None


def _estate_from_context(header: str) -> Optional[str]:
    """'ʊ Aberdeen - Winter Valley - VIC - House & Land' -> 'Aberdeen'."""
    if not header:
        return None
    first = _clean(re.sub(r"^[^A-Za-z0-9]+", "", header)).split(" - ")[0].strip()
    return first or None


# Column-name vocabulary. A header row has no price and few digits, so it used to be
# mistaken for an estate banner and remembered as context — which is how six rows ended
# up addressed "Gallery Stock List Hold Tradition Lot Land Title Date House Design ...".
_HEADER_WORDS = frozenset("""
status lot lots street st no num number land size sizes sqm m2 area floor price prices
title titled est estimated building build design designs name facade type types bed beds
bedroom bedrooms bath baths bathroom bathrooms car cars garage contract package packages
portal link links suburb estate address frontage storey storeys total totals yield rent
rental house houses furniture date dates stage release avail available availability
region state postcode deposit gross net weekly annual comments notes ref code product
completion settlement developer builder agent width depth aspect orientation
""".split())


def _is_column_header(text: str) -> bool:
    """True for a row of column names rather than a lot or an estate banner."""
    if _PRICE_ANY.search(text):
        return False
    words = [w for w in re.findall(r"[A-Za-z][A-Za-z']*", text.lower()) if len(w) > 1]
    if not words:
        return False
    known = sum(1 for w in words if w in _HEADER_WORDS)
    # a wrapped second header line is short but entirely column names ("Street # Type")
    if known == len(words) and len(words) >= 2:
        return True
    return len(words) >= 4 and known / len(words) >= 0.6


# Which column headers name money. "Land Price" is money; "Land Size" is not, and
# "Total (sqm)" is not either — hence the exclusion list, which is checked first.
_PRICE_WORD = re.compile(r"price|cost|\$|total|amount|value|deposit|rrp|package|pkg", re.I)
_NOT_PRICE_WORD = re.compile(
    r"size|area|sqm|m2|m²|width|depth|frontage|bed|bath|car|garage|storey|living|"
    r"no\.|number|lot|date|title|status|design|facade|link|brochure|yield|%|"
    r"floor|internal|external|aspect|orientation|stage|type|name|comment|note", re.I)


def _is_price_header(label: str) -> bool:
    label = label or ""
    return bool(_PRICE_WORD.search(label)) and not bool(_NOT_PRICE_WORD.search(label))


class _MoneyColumns:
    """Remembers which columns a stocklist's own header row says hold money.

    Every off-site host surfaced the same blocker: an xlsx or Google Sheets tab stores
    prices as NUMBERS, so a flattened row reads "519000.0" and `parse_fields` — which
    requires a literal "$" — throws the whole listing away. Measured on live files: Kemps
    yielded 0 listings instead of 107, AVIA 0 of 91, Invision 0 of 13, DBN Homes and
    Prosperity Residential 0 entirely, and all four Dropbox workbooks 0.

    The spreadsheet already knew which columns were money. This puts the "$" back from
    that knowledge, rather than loosening the price pattern — which would immediately
    start reading land sizes and floor areas as prices.
    """

    MIN_MONEY = 1000        # below this a bare number in a price column is not a price

    def __init__(self):
        self.cols: set = set()
        self._labels: Dict[int, str] = {}
        self._prev_was_header = False

    def feed_header(self, cells: List[Tuple[str, bool, int]], is_header: bool) -> None:
        """Header rows are often wrapped over two lines ("Land" / "Price"), so
        consecutive header rows accumulate per column instead of replacing."""
        if not is_header:
            self._prev_was_header = False
            return
        if not self._prev_was_header:
            self._labels = {}
        for text, _cur, idx in cells:
            if text:
                self._labels[idx] = f"{self._labels.get(idx, '')} {text}".strip()
        self.cols = {i for i, lab in self._labels.items() if _is_price_header(lab)}
        self._prev_was_header = True

    def render(self, cells: List[Tuple[str, bool, int]]) -> str:
        parts = []
        for text, is_currency, idx in cells:
            if not text:
                continue
            if "$" not in text and (is_currency or idx in self.cols):
                try:
                    value = float(str(text).replace(",", "").strip())
                except ValueError:
                    parts.append(text)
                    continue
                if abs(value) >= self.MIN_MONEY:
                    parts.append(f"${value:,.0f}")
                    continue
            parts.append(text)
        return _clean(" ".join(parts))


def _is_group_header(text: str, fields: Dict[str, Any]) -> bool:
    """Estate/builder/region banner rather than a lot row."""
    if fields.get("advertised_package_price"):
        return False
    if LOT_TOKEN.search(text):
        return False
    # short-ish line, mostly words, few digits
    digits = sum(c.isdigit() for c in text)
    return bool(text) and len(text) < 90 and digits <= 4


def _context_suburb(header: str) -> Optional[str]:
    """'ʊ Aberdeen - Winter Valley - VIC - House & Land' -> 'Winter Valley' (the locality part)."""
    if not header:
        return None
    cleaned = re.sub(r"^[^A-Za-z0-9]+", "", header)
    bits = [b.strip() for b in cleaned.split("-") if b.strip()]
    # drop state codes and product types
    bits = [b for b in bits if b.upper() not in ("VIC", "NSW", "QLD", "SA", "WA", "NT", "ACT", "TAS")
            and not re.search(r"house|land|terrace|townhouse|apartment|dual|key", b, re.I)]
    return bits[1] if len(bits) > 1 else (bits[0] if bits else None)


# ---------------------------------------------------------------- row iterators

def _formula_links(data: bytes) -> Dict[Tuple[int, int, int], str]:
    """(sheet index, row, column) -> url for `=HYPERLINK("...")` cells.

    Only consulted when a workbook carries no real cell hyperlinks, since loading
    it a second time for formulas is pure overhead otherwise.
    """
    out: Dict[Tuple[int, int, int], str] = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    except Exception:
        return out
    for si, sheet in enumerate(wb.worksheets):
        for row in sheet.iter_rows():
            for c in row:
                if isinstance(c.value, str) and "HYPERLINK(" in c.value.upper():
                    m = _HYPERLINK_FN.search(c.value)
                    if m:
                        out[(si, c.row, c.column)] = m.group(1)
    return out


_CURRENCY_FORMAT = re.compile(r"\$|currency|\[\$", re.I)


def _iter_rows_xlsx(data: bytes) -> Iterator[Tuple[int, List[Tuple[str, bool, int]], LinkList]]:
    """Yield (sheet index, cells, links).

    Cells are read as objects rather than values so two things survive that
    `values_only=True` discards: `cell.hyperlink` (the per-lot link) and
    `cell.number_format` (which says a bare number is money).
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    formulas = {} if any(getattr(ws, "_hyperlinks", None) for ws in wb.worksheets) \
        else _formula_links(data)
    for si, sheet in enumerate(wb.worksheets):
        for row in sheet.iter_rows():
            cells: List[Tuple[str, bool, int]] = []
            links: LinkList = []
            for c in row:
                s = str(c.value).strip() if c.value is not None else ""
                is_currency = bool(isinstance(c.value, (int, float))
                                   and _CURRENCY_FORMAT.search(c.number_format or ""))
                if s:
                    cells.append((s, is_currency, c.column))
                url = getattr(getattr(c, "hyperlink", None), "target", None) \
                    or formulas.get((si, c.row, c.column))
                if url:
                    links.append((s, str(url)))
            if cells or links:
                yield si, cells, links


def _annot_links(page: Any) -> List[Tuple[float, float, str]]:
    """(vertical centre, x0, uri) for every URI annotation on the page."""
    out: List[Tuple[float, float, str]] = []
    for a in (getattr(page, "annots", None) or []):
        uri = a.get("uri") or ((a.get("data") or {}).get("A") or {}).get("URI")
        if isinstance(uri, bytes):
            uri = uri.decode("utf-8", "ignore")
        if not uri or a.get("top") is None or a.get("bottom") is None:
            continue
        try:
            yc = (float(a["top"]) + float(a["bottom"])) / 2.0
            out.append((yc, float(a.get("x0") or 0.0), str(uri)))
        except (TypeError, ValueError):
            continue
    return out


def _links_in_band(annots: List[Tuple[float, float, str]], top: Any, bottom: Any,
                   pad: float = 2.0) -> LinkList:
    """Annotations whose centre falls on this row, ordered left to right."""
    try:
        top, bottom = float(top), float(bottom)
    except (TypeError, ValueError):
        return []
    hits = [(x0, uri) for (yc, x0, uri) in annots if top - pad <= yc <= bottom + pad]
    return [("", uri) for _, uri in sorted(hits)]


def _pdf_rows_from_tables(page: Any, annots: List[Tuple[float, float, str]]
                          ) -> List[Tuple[str, LinkList]]:
    try:
        tables = page.find_tables() or []
    except Exception:
        return []
    rows: List[Tuple[str, LinkList]] = []
    for table in tables:
        try:
            values = table.extract()
        except Exception:
            continue
        for row_obj, cells in zip(getattr(table, "rows", []), values):
            text = _row_to_text(list(cells))
            bbox = getattr(row_obj, "bbox", None)
            links = _links_in_band(annots, bbox[1], bbox[3]) if bbox else []
            if text or links:
                rows.append((text, links))
    return rows


def _pdf_rows_from_lines(page: Any, annots: List[Tuple[float, float, str]]
                         ) -> List[Tuple[str, LinkList]]:
    try:
        lines = page.extract_text_lines() or []
    except Exception:
        lines = [{"text": l, "top": None, "bottom": None}
                 for l in (page.extract_text() or "").splitlines()]
    out = []
    for ln in lines:
        text = _clean(ln.get("text", ""))
        if text:
            out.append((text, _links_in_band(annots, ln.get("top"), ln.get("bottom"))))
    return out


# Both strategies are tried on every page and the one yielding more LISTINGS wins.
# Preferring tables outright cost five builders entirely — FRD, Hudson (QLD and NSW),
# Alete and Land Build Direct publish borderless PDFs where find_tables() detects a
# handful of spurious regions, and merely producing rows was enough to suppress the
# text-line path that reads them correctly.
_PDF_STRATEGIES = (("tables", _pdf_rows_from_tables), ("lines", _pdf_rows_from_lines))


# ------------------------------------------------------------------- extraction

def _listing_from_row(text: str, links: LinkList, context: str, source_label: str,
                      builder_hint: str) -> Optional[Dict[str, Any]]:
    """A parsed listing, or None if the row is not one."""
    fields = parse_fields(text)
    # availability/storey/lot/postcode/estate/incentives, from the FULL row
    fields.update({k: v for k, v in parse_listing_features(
        text, context, fields.get("advertised_package_price")).items()
        if v is not None})
    if not fields.get("advertised_package_price"):
        return None
    # A row that is nothing but a wall of prices is a summary or a transposed layout, not a
    # listing — seen live on the commercial pages ("$1,300,000 $1,250,000 $1,250,000 ...").
    # Guarded by "and nothing that identifies a property", so the QLD dual-occupancy rows
    # (five figures, but a stock code) are unaffected.
    if len(_PRICE_ANY.findall(text)) >= 5 and not (
            fields.get("lot_number") or fields.get("bedrooms")
            or _leading_lot(text) or STREET_RE.search(text)):
        return None
    fields["source_text"] = text                 # UNtruncated: the parser needs it all
    if not fields.get("estate_name"):
        fields["estate_name"] = _estate_from_context(context)
    label = _address_label(text, fields, context)
    if label:                                    # beats the raw row parse_fields found
        fields["lot_address"] = label
    elif not fields.get("lot_address"):
        fields["lot_address"] = _clean(text)[:110]
    if not fields.get("suburb"):
        fields["suburb"] = _context_suburb(context)
    fields.update(_link_fields(links))
    filled = sum(1 for k in ("bedrooms", "land_size_sqm", "suburb", "house_size_sqm")
                 if fields.get(k))
    return {
        **fields,
        "builder_name": builder_hint,
        "estate_context": _clean(context)[:110],
        "source_url_or_ref": source_label,
        "extraction_confidence": round(min(1.0, 0.5 + 0.12 * filled), 2),
    }


def _rows_to_listings(rows: List[Tuple[str, LinkList]], context: str, source_label: str,
                      builder_hint: str) -> Tuple[List[Dict[str, Any]], str]:
    """Parse a batch of rows, carrying the estate/group header context through them.

    Returns (listings, context) rather than mutating, so two competing PDF strategies
    can each be tried on a page without one polluting the other's header context.
    """
    out: List[Dict[str, Any]] = []
    for text, links in rows:
        if not text:
            continue
        if _is_column_header(text):
            continue                         # column names: neither a lot nor a banner
        listing = _listing_from_row(text, links, context, source_label, builder_hint)
        if listing is None:
            if _is_group_header(text, {}):   # no price: _listing_from_row already told us
                context = text
            continue
        out.append(listing)
    return out, context


def _assign_variant_ordinals(rows: List[Dict[str, Any]]) -> int:
    """Number rows that share an identity but are genuinely different packages.

    Identity ignores price so a price move updates in place. The cost is that two rows
    in one file which differ ONLY by price collapse into one — verified on the APLACE,
    Met Invest and Paramount stocklists. Numbering the siblings keeps both.

    Byte-identical rows are left alone: those are duplicates in the source file (a
    table detected twice across a page break) and SHOULD collapse. Returns the number
    of siblings numbered, so a caller can report it.
    """
    from database import building_content_hash          # local: avoids an import cycle

    groups: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        groups.setdefault(building_content_hash(r), []).append(r)
    numbered = 0
    for group in groups.values():
        if len(group) < 2:
            continue
        seen, n = set(), 0
        for r in group:
            sig = (str(r.get("source_text")), r.get("advertised_package_price"))
            if sig in seen:
                continue                    # true duplicate — let it collapse
            seen.add(sig)
            if n:
                r["variant_ordinal"] = n
                numbered += 1
            n += 1
    return numbered


def _log_yield(kind: str, source_label: str, out: List[Dict[str, Any]]) -> None:
    numbered = _assign_variant_ordinals(out)
    links = sum(1 for o in out if o.get("listing_url") or o.get("floorplan_url")
                or o.get("brochure_url"))
    logger.info("%s %s -> %d listing(s), %d with a per-lot link%s", kind,
                source_label or "(file)", len(out), links,
                f", {numbered} price-only sibling(s) numbered" if numbered else "")


def extract_from_xlsx(data: bytes, source_label: str = "", builder_hint: str = "") -> List[Dict[str, Any]]:
    if not XLSX_AVAILABLE:
        logger.error("openpyxl not installed — cannot read xlsx stocklists.")
        return []
    try:
        rows = list(_iter_rows_xlsx(data))
    except Exception as e:
        logger.warning("could not open xlsx %s: %s", source_label, e)
        return []

    out: List[Dict[str, Any]] = []
    context, sheet, money = "", -1, _MoneyColumns()
    for si, cells, links in rows:
        if si != sheet:
            context, sheet, money = "", si, _MoneyColumns()   # headers are per sheet
        plain = _clean(" ".join(t for t, _c, _i in cells))
        if not plain:
            continue
        is_header = _is_column_header(plain)
        money.feed_header(cells, is_header)
        if is_header:
            continue
        listings, context = _rows_to_listings([(money.render(cells), links)], context,
                                              source_label, builder_hint)
        out.extend(listings)
    _log_yield("xlsx", source_label, out)
    return out


def extract_from_pdf(data: bytes, source_label: str = "", builder_hint: str = "") -> List[Dict[str, Any]]:
    if not PDF_AVAILABLE:
        logger.error("pdfplumber not installed — cannot read pdf stocklists.")
        return []
    out: List[Dict[str, Any]] = []
    used = {}
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            context = ""
            for page in pdf.pages[:25]:
                annots = _annot_links(page)
                best, best_ctx, best_name = [], context, ""
                for name, strategy in _PDF_STRATEGIES:
                    listings, ctx = _rows_to_listings(strategy(page, annots), context,
                                                      source_label, builder_hint)
                    if len(listings) > len(best):
                        best, best_ctx, best_name = listings, ctx, name
                out.extend(best)
                context = best_ctx
                if best_name:
                    used[best_name] = used.get(best_name, 0) + len(best)
    except Exception as e:
        logger.warning("could not read pdf %s: %s", source_label, e)
    if len(used) > 1:
        logger.info("pdf %s: mixed page layouts — %s", source_label or "(file)",
                    ", ".join(f"{k} {v}" for k, v in used.items()))
    _log_yield("pdf", source_label, out)
    return out


def extract_from_csv(data: bytes, source_label: str = "", builder_hint: str = "") -> List[Dict[str, Any]]:
    """CSV/TSV stocklists — what a Google Sheets or Smartsheet tab exports as."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        logger.warning("could not decode csv %s", source_label)
        return []
    sample = text[:4000]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    rows: List[Tuple[str, LinkList]] = []
    money = _MoneyColumns()
    for raw in csv.reader(io.StringIO(text), delimiter=delimiter):
        cells = [(str(v).strip(), False, i) for i, v in enumerate(raw) if str(v).strip()]
        plain = _clean(" ".join(t for t, _c, _i in cells))
        if not plain:
            continue
        is_header = _is_column_header(plain)
        money.feed_header(cells, is_header)
        if not is_header:
            rows.append((money.render(cells), []))
    out, _ = _rows_to_listings(rows, "", source_label, builder_hint)
    _log_yield("csv", source_label, out)
    return out


def extract_stocklist(data: bytes, source_label: str = "", builder_hint: str = "") -> List[Dict[str, Any]]:
    """Dispatch on content. Magic bytes for the binary formats, then a text sniff — a
    Google Sheets or Smartsheet export arrives as CSV with no magic bytes at all."""
    if not data:
        return []
    if data[:2] == b"PK":
        return extract_from_xlsx(data, source_label, builder_hint)
    if data[:4] == b"%PDF":
        return extract_from_pdf(data, source_label, builder_hint)
    head = data[:600].lstrip()
    if head[:1] in (b"<",):
        # An HTML page where a file was expected is a sign-in wall or an error page,
        # never stock. Say so plainly rather than parsing a login form.
        logger.warning("stocklist %s returned an HTML page, not a file — the link most "
                       "likely needs a signed-in session.", source_label)
        return []
    if b"," in head or b"\t" in head:
        return extract_from_csv(data, source_label, builder_hint)
    logger.warning("unrecognised stocklist format for %s (starts %r)", source_label, data[:12])
    return []
