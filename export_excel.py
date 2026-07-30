"""
Export the harvested stock as a formatted Excel workbook.

Two sheets:
  Stock    every listing, in the column order Coleen read the sheet in on 29 July,
           with the per-lot PDF / floor plan / brochure as clickable links
  Summary  coverage figures, written as FORMULAS over the Stock sheet so they
           recalculate if anyone filters, edits or adds rows

Blanks are deliberate throughout and the Summary sheet says so: this project never
guesses a builder, a state or a bedroom count, because a wrong one in front of a buyer
is worse than an obvious gap.

Usage:
    python export_excel.py                 # output/spb_stock_<date>.xlsx
    python export_excel.py --out FILE.xlsx
"""

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from database import ResearchDatabase
from export_csv import BUILDING_COLS

FONT = "Arial"
STAMP = datetime.now().strftime("%Y%m%d")

# (db key -> excel number format). Everything else is left as text.
MONEY = {"price", "land_price", "build_price", "incentive_amount", "benchmark_median"}
SIZE = {"land_sqm", "house_sqm", "frontage_m"}
COUNT = {"bedrooms", "bathrooms", "car_spaces"}
LINKS = {"listing_url": "open", "floorplan_url": "plan", "brochure_url": "file"}

# Availability, coloured the way the dashboard colours it.
AVAIL_FILL = {
    "available": PatternFill("solid", fgColor="E8F5EE"),
    "under offer": PatternFill("solid", fgColor="FEF6E7"),
    "on hold": PatternFill("solid", fgColor="FEF6E7"),
    "reserved": PatternFill("solid", fgColor="FEF6E7"),
    "sold": PatternFill("solid", fgColor="FDECEA"),
    "not available": PatternFill("solid", fgColor="FDECEA"),
    "leased": PatternFill("solid", fgColor="FDECEA"),
}
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
BLANK_FILL = PatternFill("solid", fgColor="F1F5F9")

WIDTHS = {  # by header label, so it survives a reordering of BUILDING_COLS
    "Builder / Development": 30, "Address": 34, "Suburb": 18, "State": 7,
    "Availability": 13, "Storey": 8, "Package Price": 14, "Land Size m2": 11,
    "Land Price": 13, "House Size m2": 12, "House Price": 13, "Beds": 6,
    "Baths": 6, "Cars": 6, "Frontage m": 10, "Estate": 20,
    "Title / Registration": 18, "Incentive $": 11, "Incentive": 26,
    "Market Median": 13, "Variance vs Market %": 17, "Vs Market": 15,
    "PDF / Listing": 12, "Floor Plan": 11, "Brochure": 10, "Date Checked": 12,
    "Source": 24, "Attribution": 13, "State From": 26, "Builder From": 26,
    "Product": 13, "Confidence": 10, "Source File": 40,
}


def _rows(db):
    rows = db.get_buildings()
    # Same order as the CSV: state, then builder with blanks LAST, then suburb, price.
    rows.sort(key=lambda r: (
        r.get("state") or "zz",
        not (r.get("builder_name") or "").strip(),
        (r.get("builder_name") or "").lower(),
        (r.get("suburb") or "").lower(),
        r.get("price") or 0,
    ))
    return rows


def _write_stock(ws, rows):
    labels = [label for _k, label in BUILDING_COLS]
    ws.append(labels)
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for r in rows:
        excel_row = []
        for key, _label in BUILDING_COLS:
            v = r.get(key)
            if isinstance(v, str):
                v = " ".join(v.split())
            excel_row.append(v)
        ws.append(excel_row)

    for i, (key, label) in enumerate(BUILDING_COLS, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = WIDTHS.get(label, 14)
        fmt = ("$#,##0" if key in MONEY else "0.0" if key in SIZE
               else "0.#" if key in COUNT
               else '0.0"%"' if key == "benchmark_variance_pct" else None)
        for cell in ws[col][1:]:
            cell.font = Font(name=FONT, size=10)
            if fmt:
                cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right")

    # links, and the two columns whose blanks mean something
    keys = [k for k, _l in BUILDING_COLS]
    for key, text in LINKS.items():
        if key not in keys:
            continue
        col = get_column_letter(keys.index(key) + 1)
        for cell in ws[col][1:]:
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.value = text
                cell.font = Font(name=FONT, size=10, color="1E40AF", underline="single")
            cell.alignment = Alignment(horizontal="center")

    for key in ("builder_name", "state"):
        col = get_column_letter(keys.index(key) + 1)
        for cell in ws[col][1:]:
            if not cell.value:
                cell.fill = BLANK_FILL              # a gap that is meant to be visible

    col = get_column_letter(keys.index("availability_status") + 1)
    for cell in ws[col][1:]:
        fill = AVAIL_FILL.get(str(cell.value or "").strip().lower())
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    last = ws.max_row
    ws.freeze_panes = "C2"                          # keep builder + address in view
    ws.auto_filter.ref = f"A1:{get_column_letter(len(BUILDING_COLS))}{last}"
    return last


def _write_summary(ws, last_row, generated):
    keys = [k for k, _l in BUILDING_COLS]
    C = {k: get_column_letter(keys.index(k) + 1) for k in keys}

    def rng(key):
        return f"Stock!${C[key]}$2:${C[key]}${last_row}"

    ws["A1"] = "Smart Property Buying — harvested stock"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = f"Generated {generated} from the live harvest database."
    ws["A2"].font = Font(name=FONT, size=10, italic=True, color="64748B")

    # Every figure is a formula over the Stock sheet, so filtering or adding rows keeps
    # this honest rather than freezing a number that was true once.
    metrics = [
        ("Listings", f"=COUNTA({rng('lot_address')})"),
        ("Builders / developments named",
         f"=SUMPRODUCT(({rng('builder_name')}<>\"\")/COUNTIF({rng('builder_name')},"
         f"{rng('builder_name')}&\"\"))"),
        ("With a named builder", f'=COUNTIF({rng("builder_name")},"?*")'),
        ("With a state", f'=COUNTIF({rng("state")},"?*")'),
        ("With availability", f'=COUNTIF({rng("availability_status")},"?*")'),
        ("With a suburb", f'=COUNTIF({rng("suburb")},"?*")'),
        ("With a bedroom count", f"=COUNT({rng('bedrooms')})"),
        ("With land size", f"=COUNT({rng('land_sqm')})"),
        ("With house size", f"=COUNT({rng('house_sqm')})"),
        ("With a PDF, plan or brochure",
         f'=SUMPRODUCT(--(({rng("listing_url")}<>"")+({rng("floorplan_url")}<>"")'
         f'+({rng("brochure_url")}<>"")>0))'),
        ("Total package value", f"=SUM({rng('price')})"),
        ("Median package price", f"=MEDIAN({rng('price')})"),
    ]
    row = 4
    ws[f"A{row}"] = "Coverage"
    ws[f"A{row}"].font = Font(name=FONT, bold=True, size=11)
    row += 1
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        ws[f"A{row}"].font = Font(name=FONT, size=10)
        ws[f"B{row}"].font = Font(name=FONT, size=10, bold=True)
        if "package" in label.lower() and "With" not in label:
            ws[f"B{row}"].number_format = "$#,##0"
        else:
            ws[f"B{row}"].number_format = "#,##0"
        row += 1

    row += 1
    ws[f"A{row}"] = "By state"
    ws[f"A{row}"].font = Font(name=FONT, bold=True, size=11)
    row += 1
    for st in ("VIC", "NSW", "QLD", "SA", "WA", "TAS", "NT", "ACT"):
        ws[f"A{row}"] = st
        ws[f"B{row}"] = f'=COUNTIF({rng("state")},"{st}")'
        ws[f"A{row}"].font = Font(name=FONT, size=10)
        ws[f"B{row}"].font = Font(name=FONT, size=10)
        ws[f"B{row}"].number_format = "#,##0"
        row += 1
    ws[f"A{row}"] = "not established"
    ws[f"B{row}"] = f'=COUNTA({rng("lot_address")})-COUNTIF({rng("state")},"?*")'
    ws[f"A{row}"].font = Font(name=FONT, size=10, italic=True)
    ws[f"B{row}"].font = Font(name=FONT, size=10, italic=True)
    row += 2

    ws[f"A{row}"] = "By availability"
    ws[f"A{row}"].font = Font(name=FONT, bold=True, size=11)
    row += 1
    for label in ("Available", "Under Offer", "On Hold", "Reserved", "Sold",
                  "Not Available", "Leased"):
        ws[f"A{row}"] = label
        ws[f"B{row}"] = f'=COUNTIF({rng("availability_status")},"{label}")'
        ws[f"A{row}"].font = Font(name=FONT, size=10)
        ws[f"B{row}"].font = Font(name=FONT, size=10)
        ws[f"B{row}"].number_format = "#,##0"
        row += 1

    row += 1
    ws[f"A{row}"] = "Reading the blanks"
    ws[f"A{row}"].font = Font(name=FONT, bold=True, size=11)
    row += 1
    for note in (
        "Shaded cells in Builder and State are BLANK ON PURPOSE. Nothing here is guessed: a "
        "wrong builder or state in front of a buyer is worse than a visible gap.",
        "Builder blank, Attribution 'state_pooled' — the lot came from one of E-Agent's pooled "
        "state stocklists, which list several builders in one file without saying which lot is whose.",
        "Builder blank, Attribution 'project' — an apartment or townhouse development. E-Agent "
        "lists these by development rather than builder, and this one is not named on the page or "
        "in its price list.",
        "'State From' records which signal decided each state — the listing's own postcode, its "
        "suburb, or the E-Agent page it came from. 'conflicting signals' means two disagreed.",
        "Beds, baths, cars, land and house size are sparse because most builders' stocklists put "
        "them in columns the extractor does not yet map. That work is in progress.",
    ):
        ws[f"A{row}"] = note
        ws[f"A{row}"].font = Font(name=FONT, size=9, color="475569")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{row}:F{row}")
        ws.row_dimensions[row].height = 28
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16


def main(out: Path = None) -> Path:
    db = ResearchDatabase()
    rows = _rows(db)
    generated = datetime.now().strftime("%d %b %Y, %H:%M")

    wb = openpyxl.Workbook()
    # Buildings FIRST and selected. With Summary first the file opened on a page of
    # totals and looked as though it held no stock at all.
    stock = wb.active
    stock.title = "Buildings"
    summary = wb.create_sheet("Summary")
    last = _write_stock(stock, rows)
    _write_summary(summary, last, generated)
    wb.active = 0
    stock.sheet_view.tabSelected = True
    summary.sheet_view.tabSelected = False

    # openpyxl writes formulas with no cached value, so the Summary sheet would read as
    # blank until something calculates it. This makes Excel (and Sheets, and LibreOffice)
    # recalculate the moment the file is opened.
    wb.calculation.fullCalcOnLoad = True

    config.OUTPUT_DIR.mkdir(exist_ok=True)
    out = out or config.OUTPUT_DIR / f"spb_stock_{STAMP}.xlsx"
    wb.save(out)
    print(f"[OK] {out}")
    print(f"     Stock:   {len(rows):,} listings x {len(BUILDING_COLS)} columns")
    print(f"     Summary: coverage figures as formulas over the Stock sheet")
    return out


if __name__ == "__main__":
    dest = None
    if "--out" in sys.argv:
        dest = Path(sys.argv[sys.argv.index("--out") + 1])
    main(dest)
