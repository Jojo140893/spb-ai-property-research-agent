"""
Real market medians from official open data — free, licensed, and automatable.

WHY THIS EXISTS. Coleen asked on 3 Aug for benchmarking, not a filter, and named
realestate.com.au. That door is bolted: REA answers the FIRST unauthenticated request
with HTTP 429 and a Kasada bot-detection challenge, and domain.com.au answers 403. The
only way past either is to defeat a bot-detection system, which is not something this
project will build. REA sells the same data as PropTrack precisely because the front
door is shut.

So until a licence is bought, this is the market basis that actually works today:
government property-sales data, published under Creative Commons Attribution, fetched
through data.gov.au's catalogue API. No key, no bot wall, no terms to breach — the
licence explicitly permits reuse with attribution, which the app records per figure.

    python market_open_data.py            # refresh the cache and report coverage
    python market_open_data.py --status   # what is cached, how old, what it covers

WHAT IT IS AND IS NOT. These are MEDIAN SALE PRICES BY SUBURB — actual sold prices,
which for judging whether a package is well bought are better evidence than the asking
prices on a portal. What they are not is individual listings, so they answer Benchmark B
("is this a good buy for this suburb?") and NOT Benchmark A ("is there a specific
cheaper property a buyer could find right now"). Benchmark A still needs a listings
feed. Saying which of the two a number supports is the whole reason benchmark_basis
exists, and this never claims the other.

COVERAGE IS PARTIAL AND SAID OUT LOUD. Victoria publishes suburb medians in this form;
at the time of writing it covers 62 of our 87 located Victorian suburbs (71%). Other
states publish on their own schedules and formats, and each needs its own reader. A
suburb with no median gets no market claim rather than a borrowed one.
"""

import argparse
import json
import re
import sys
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Optional, Tuple

sys.path.insert(0, str(Path(__file__).resolve().parent))

import config                                                        # noqa: E402

CACHE_DIR = Path(__file__).resolve().parent / "data" / "market"
CKAN = "https://data.gov.au/data/api/3/action/package_search"
UA = ("SPB-PropertyResearch/1.0 (+market comparables for client reporting; "
      "contact support@csmsynergy.com)")

# Refetched only when stale: these are quarterly and yearly series, so a daily download
# would be pure noise against someone else's bandwidth.
MAX_AGE_DAYS = 14
TIMEOUT = 90

# One entry per published series. `state` is what the figures describe; `dwelling` keeps
# houses and units apart, because a unit median against a house-and-land package is the
# same category error as a land price against a package.
SERIES = (
    {"key": "vic_house", "state": "VIC", "dwelling": "house",
     "query": "Victorian Property Sales Report Median House by Suburb Time Series",
     "match": re.compile(r"houses?-by-suburb.*\.xlsx?$", re.I)},
    {"key": "vic_unit", "state": "VIC", "dwelling": "unit",
     "query": "Victorian Property Sales Report Median Unit by Suburb Time Series",
     "match": re.compile(r"units?-by-suburb.*\.xlsx?$", re.I)},
)


def _get(url: str, timeout: int = TIMEOUT) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def discover(series: Dict) -> Optional[Dict[str, str]]:
    """The newest matching resource for a series: {url, title, licence, publisher}.

    Resolved through the catalogue rather than hardcoded, because the file name carries
    the year ('houses-by-suburb-2014-2024.xlsx') and a pinned URL silently goes stale
    the moment the next one is published — which is exactly the kind of quiet staleness
    this codebase keeps having to fix.
    """
    url = f"{CKAN}?q={urllib.parse.quote(series['query'])}&rows=8"
    try:
        found = json.loads(_get(url, timeout=40).decode("utf-8", "replace"))
    except Exception as exc:                                          # noqa: BLE001
        print(f"  [!] catalogue unreachable for {series['key']}: {exc}")
        return None
    for pkg in found.get("result", {}).get("results", []):
        for res in pkg.get("resources", []):
            href = (res.get("url") or "").strip()
            # web.archive.org copies appear in the catalogue for withdrawn quarters;
            # they are snapshots of old data, not the current series.
            if "web.archive.org" in href or not series["match"].search(href):
                continue
            return {"url": href, "title": pkg.get("title", ""),
                    "licence": pkg.get("license_title", ""),
                    "publisher": (pkg.get("organization") or {}).get("title", "")}
    return None


def refresh(force: bool = False) -> Dict[str, Dict]:
    """Download each series if the cache is stale. Returns what is now available."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    out = {}
    for series in SERIES:
        blob = CACHE_DIR / f"{series['key']}.xlsx"
        meta_path = CACHE_DIR / f"{series['key']}.json"
        fresh = (blob.exists() and not force
                 and datetime.fromtimestamp(blob.stat().st_mtime)
                 > datetime.now() - timedelta(days=MAX_AGE_DAYS))
        if fresh:
            meta = json.loads(meta_path.read_text(encoding="utf-8")) if meta_path.exists() else {}
            print(f"  {series['key']:<10} cached ({datetime.fromtimestamp(blob.stat().st_mtime):%d %b})")
            out[series["key"]] = {**series, **meta}
            continue
        found = discover(series)
        if not found:
            print(f"  {series['key']:<10} NOT FOUND in the catalogue — left as it was")
            if blob.exists():
                out[series["key"]] = {**series}
            continue
        try:
            blob.write_bytes(_get(found["url"]))
        except Exception as exc:                                      # noqa: BLE001
            print(f"  {series['key']:<10} download failed: {exc}")
            continue
        meta_path.write_text(json.dumps(found, indent=1), encoding="utf-8")
        print(f"  {series['key']:<10} downloaded {blob.stat().st_size:,} bytes "
              f"— {found['licence'] or 'licence not stated'}")
        out[series["key"]] = {**series, **found}
    return out


def _read_medians(path: Path) -> Tuple[Dict[str, float], str]:
    """{suburb_lower: median} for the most recent column that has figures, and its year."""
    try:
        import openpyxl
    except ImportError:
        print("  [!] openpyxl missing — run: pip install openpyxl")
        return {}, ""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = next((r for r in rows if r and str(r[0] or "").strip().lower() == "locality"), None)
    if not header:
        return {}, ""
    years = {i: str(v).strip() for i, v in enumerate(header)
             if v and str(v).strip().isdigit()}
    if not years:
        return {}, ""
    # The newest column that actually holds numbers. The last column is sometimes a
    # part-year with '-' in most rows, and taking it blindly would report a median for a
    # handful of suburbs and nothing for the rest.
    best, best_year = {}, ""
    for idx in sorted(years, reverse=True):
        found = {}
        for r in rows:
            name = str(r[0] or "").strip()
            if not name or name.lower() == "locality" or idx >= len(r):
                continue
            try:
                found[name.lower()] = float(str(r[idx]).replace(",", "").replace("$", ""))
            except (TypeError, ValueError):
                continue
        if len(found) > len(best):
            best, best_year = found, years[idx]
        if len(found) >= 200:            # a full year's coverage; no need to look older
            break
    return best, best_year


class OpenDataMarket:
    """Suburb medians from official open data. Answers only what it can evidence."""

    name = "open-data"

    def __init__(self, refresh_if_stale: bool = False):
        self.series: Dict[str, Dict] = {}
        self.medians: Dict[Tuple[str, str, str], float] = {}   # (suburb, state, dwelling)
        self.year: Dict[str, str] = {}
        if refresh_if_stale:
            refresh()
        self._load()

    def _load(self):
        for series in SERIES:
            blob = CACHE_DIR / f"{series['key']}.xlsx"
            meta_path = CACHE_DIR / f"{series['key']}.json"
            if not blob.exists():
                continue
            meds, year = _read_medians(blob)
            if not meds:
                continue
            self.year[series["key"]] = year
            meta = json.loads(meta_path.read_text(encoding="utf-8")) if meta_path.exists() else {}
            self.series[series["key"]] = {**series, **meta}
            for suburb, value in meds.items():
                self.medians[(suburb, series["state"], series["dwelling"])] = value

    @property
    def loaded(self) -> bool:
        return bool(self.medians)

    def median(self, suburb: str, state: str, dwelling: str = "house"
               ) -> Optional[Tuple[float, str]]:
        """(median, attribution) or None. Never a borrowed figure from another suburb."""
        key = ((suburb or "").strip().lower(), (state or "").strip().upper(), dwelling)
        value = self.medians.get(key)
        if not value:
            return None
        series = next((s for s in self.series.values()
                       if s["state"] == key[1] and s["dwelling"] == dwelling), {})
        year = self.year.get(series.get("key", ""), "")
        who = series.get("publisher") or "state government open data"
        licence = series.get("licence") or "open licence"
        return value, (f"{dwelling} median for {suburb.title()} {key[1]}, {year} — "
                       f"{who} ({licence}), via data.gov.au")


def _status() -> int:
    m = OpenDataMarket()
    print("=" * 70)
    print("  Open-data market medians")
    print("=" * 70)
    if not m.loaded:
        print("  nothing cached yet — run: python market_open_data.py")
        return 1
    for key, series in m.series.items():
        n = sum(1 for k in m.medians if k[1] == series["state"] and k[2] == series["dwelling"])
        print(f"  {key:<10} {n:>5} suburb(s)  {series['state']} {series['dwelling']}s, "
              f"{m.year.get(key, '?')}")
        print(f"             {series.get('licence', 'licence not stated')}")
    # What it covers of our own stock — the number that decides whether it is useful.
    try:
        import sqlite3

        import suburb_quality
        conn = sqlite3.connect(str(config.DATABASE_PATH))
        conn.row_factory = sqlite3.Row
        ours = {}
        for r in conn.execute(
                "SELECT suburb, state, lot_address, street_address, source_text "
                "FROM buildings WHERE price > 0 "
                "  AND (superseded_by IS NULL OR superseded_by = '')"):
            row = dict(r)
            name, why = suburb_quality.resolve(row)
            if suburb_quality.is_located(why):
                ours[(name.lower(), (row["state"] or "").upper())] = \
                    ours.get((name.lower(), (row["state"] or "").upper()), 0) + 1
        covered = {k: n for k, n in ours.items()
                   if any((k[0], k[1], d) in m.medians for d in ("house", "unit"))}
        listings = sum(ours.values())
        print(f"\n  our located stock : {listings:,} listing(s) across {len(ours)} suburb(s)")
        print(f"  with a median     : {sum(covered.values()):,} listing(s) across "
              f"{len(covered)} suburb(s)")
        gap = sorted(((n, k) for k, n in ours.items() if k not in covered), reverse=True)[:5]
        if gap:
            print("  biggest gaps      : "
                  + ", ".join(f"{k[0].title()} {k[1]} ({n})" for n, k in gap))
    except Exception as exc:                                          # noqa: BLE001
        print(f"\n  (could not measure coverage: {exc})")
    return 0


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--status", action="store_true", help="report only, fetch nothing")
    ap.add_argument("--force", action="store_true", help="refetch even if cached")
    args = ap.parse_args(argv)
    if args.status:
        return _status()
    print("=" * 70)
    print("  Refreshing open-data market medians")
    print("=" * 70)
    refresh(force=args.force)
    print()
    return _status()


if __name__ == "__main__":
    sys.exit(main())
