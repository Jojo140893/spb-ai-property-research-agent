"""
Tests for per-listing links in stocklists.

Coleen: "if someone wants to click, they can click here and see the PDF."

Today every row from a stocklist file shares that file's URL, so the export has no
per-lot link at all. The link exists in the source — as an XLSX cell hyperlink
(often on a bare "Download" cell, which is why the text alone was useless) or as a
PDF page annotation. These tests build both file types and assert the link reaches
the listing, and lands in the right column.
"""

import io

import openpyxl

from sources.spreadsheet_extract import (_annot_links, _classify_link,
                                         _link_fields, _links_in_band,
                                         extract_from_xlsx)

FLYER = "https://e-agent.com.au/_files/ugd/lot82-package.pdf"
PLAN = "https://e-agent.com.au/_files/ugd/lot82-floorplan.pdf"
# Row shape taken from the live "VIC Regional" stocklist.
_ROW = ["Lot 82 Aberdeen", 282, 142.8, 12.0, "Sep-26", "Empley 15", "3x2x2",
        "$205,000", "$335,220", "$540,220", "Available"]


def _workbook(hyperlink=True) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Aberdeen - Winter Valley - VIC - House & Land"])
    ws.append(_ROW + ["Download", "Floorplan"])
    if hyperlink:
        ws.cell(row=2, column=12).hyperlink = FLYER
        ws.cell(row=2, column=13).hyperlink = PLAN
    else:                                   # the =HYPERLINK() form, whose target is
        ws.cell(row=2, column=12).value = f'=HYPERLINK("{FLYER}","Download")'
        ws.cell(row=2, column=13).value = f'=HYPERLINK("{PLAN}","Floorplan")'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_cell_hyperlink_reaches_the_listing():
    rows = extract_from_xlsx(_workbook(), source_label="stock.xlsx")
    assert len(rows) == 1, f"expected 1 listing, got {len(rows)}"
    r = rows[0]
    assert r["listing_url"] == FLYER, f"per-lot link lost: {r.get('listing_url')}"
    assert r["floorplan_url"] == PLAN, f"floorplan misfiled: {r.get('floorplan_url')}"
    # and the row still parses as it did before
    assert r["advertised_package_price"] == 540_220
    assert r["lot_address"] == "Lot 82, Aberdeen"


def test_hyperlink_formula_target_is_recovered():
    """data_only=True returns the cached text of a formula, discarding its URL."""
    rows = extract_from_xlsx(_workbook(hyperlink=False), source_label="stock.xlsx")
    assert len(rows) == 1
    assert rows[0]["listing_url"] == FLYER
    assert rows[0]["floorplan_url"] == PLAN


def test_rows_without_links_are_unaffected():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_ROW)
    buf = io.BytesIO()
    wb.save(buf)
    rows = extract_from_xlsx(buf.getvalue(), source_label="stock.xlsx")
    assert len(rows) == 1
    assert rows[0].get("listing_url") is None
    assert rows[0]["advertised_package_price"] == 540_220


def test_link_kind_is_taken_from_anchor_or_url():
    assert _classify_link("https://x/a.pdf", "Download") == "listing_url"
    assert _classify_link("https://x/lot-9-floorplan.pdf", "") == "floorplan_url"
    assert _classify_link("https://x/a.pdf", "Floor Plan") == "floorplan_url"
    assert _classify_link("https://x/EstateBrochure.pdf", "") == "brochure_url"
    assert _classify_link("https://x/marketing-flyer.pdf", "") == "brochure_url"
    # first of each kind wins; a second generic link does not overwrite the first
    got = _link_fields([("Download", "https://x/1.pdf"), ("More", "https://x/2.pdf"),
                        ("Plan", "https://x/lot9-floorplan.pdf")])
    assert got == {"listing_url": "https://x/1.pdf",
                   "floorplan_url": "https://x/lot9-floorplan.pdf"}, got


def test_pdf_annotations_are_matched_to_their_row_by_position():
    """A PDF stocklist's links are page annotations with no row of their own — they
    are only usable if matched to the row they sit on."""
    class _Page:
        annots = [
            {"uri": FLYER, "top": 100.0, "bottom": 110.0, "x0": 400.0},
            {"uri": PLAN, "top": 101.0, "bottom": 109.0, "x0": 480.0},
            {"uri": "https://x/other.pdf", "top": 300.0, "bottom": 310.0, "x0": 400.0},
            {"uri": None, "top": 100.0, "bottom": 110.0, "x0": 10.0},      # not a link
            {"uri": FLYER, "top": None, "bottom": None, "x0": 10.0},       # no position
        ]
    annots = _annot_links(_Page())
    assert len(annots) == 3, f"annotation parsing wrong: {annots}"
    on_row = _links_in_band(annots, 98.0, 112.0)
    assert [u for _, u in on_row] == [FLYER, PLAN], on_row   # left to right
    assert _links_in_band(annots, 200.0, 220.0) == []        # nothing on an empty band
    assert _link_fields(on_row) == {"listing_url": FLYER, "floorplan_url": PLAN}


def test_address_column_is_a_label_not_the_whole_row():
    """The bug Coleen pointed at on screen. `parse_fields` sets lot_address to the
    whole line containing "Lot N", which for a flattened stocklist row is the entire
    row — so the stocklist path must override it."""
    from sources.spreadsheet_extract import _address_label
    from sources.adaptive_extract import parse_fields
    from sources.feature_extract import parse_listing_features

    def label(text, ctx=""):
        f = parse_fields(text)
        f.update({k: v for k, v in parse_listing_features(
            text, ctx, f.get("advertised_package_price")).items() if v is not None})
        return _address_label(text, f, ctx)

    assert label(" ".join(str(c) for c in _ROW),
                 "Aberdeen - Winter Valley - VIC - House & Land") == "Lot 82, Aberdeen"
    # a real street address must win over the lot number and survive intact
    assert label("Available 25/03/2026 105 Almond Street $890,000",
                 "Denman - Highfields Estate") == "105 Almond Street, Highfields"
    # stock codes are already labels; don't prefix them with "Lot"
    assert label("CC-0114 506 Titled 200 5 + 5 + 3 MORETON $ 839,000 $ 694,199 "
                 "$ 1,533,199") == "CC-0114"
    # nothing recognisable -> None, so the caller keeps whatever it already had
    assert label("Package from $650,000 enquire now") is None

    # Rows from the per-builder crawl. Several builders' files ARE just an address, so
    # composing a label from parts would only throw half of it away.
    assert label("Lot 444, Hillcrest $1,353,596") == "Lot 444, Hillcrest"
    assert label("4 Windsor Street, BRISBANE NORTH $980,780") == "4 Windsor Street, BRISBANE NORTH"
    assert label("520 Stony Drive, Alluvium $680,294") == "520 Stony Drive, Alluvium"
    assert label("1368 Margery St, Toolern Waters , Melton South 3338 "
                 "$596,000") == "1368 Margery St, Toolern Waters"
    # a bare leading number is the lot on several stocklists ("103 Samara Estate ...")
    assert label("103 Samara Estate Fraser Rise Available Bristol 15 $350,000 $335,970 "
                 "$685,970") == "Lot 103, Samara"
    # ...but Gallery Group's sheet title is not a locality
    assert label("100 Park Royal Crescent, GALLERY STOCK LIST 2026 "
                 "$1,328,810") == "100 Park Royal Crescent"
    # a grid of figures is not an address, even when it is short
    assert label("12 Brittlewood 300 26 CTM 5 Yes 3 2 N/A Double $1,430,000") == "Lot 12"


def test_column_header_rows_never_become_estate_context():
    """A header row has no price and few digits, so it used to pass for an estate banner
    and be remembered as context — which is how six live rows ended up addressed
    "Gallery Stock List Hold Tradition Lot Land Title Date House Design ...".
    All fixtures below are real header lines from the builders' own files."""
    from sources.spreadsheet_extract import _is_column_header

    for header in (
        "Status Lot/ Land Size Land Price Est. Title Building Design Name Floor Area "
        "Bed / Bath / Car Build Package Price Est. Gross",
        "Gallery Stock List Hold Tradition Lot Land Title Date House Design Facade "
        "Land Price House Price Furniture Total",
        "Status,Lot,Estate,Suburb,Land,House,Design,Beds,Land Price,Build Price,Package",
        "Street # Type",                      # a wrapped second header line
    ):
        assert _is_column_header(header), f"header not recognised: {header[:60]}"

    for not_a_header in (
        "Aberdeen - Winter Valley - VIC - House & Land",     # estate banner
        "ANGLE VALE SA - 23 - NORTH ESTATE, Angle Vale",
        "Paradise in Parkinson - Stage 2, Parkinson",
        "Lot 82 Aberdeen 282 142.8 12.0 Sep-26 Empley 15 3x2x2 $205,000 $335,220 "
        "$540,220 Available",                                # a real lot
    ):
        assert not _is_column_header(not_a_header), f"wrongly a header: {not_a_header[:60]}"


def test_csv_stocklist_is_parsed():
    """A Google Sheets or Smartsheet tab exports as CSV, which has no magic bytes at all
    — the dispatcher used to reject it as an unrecognised format."""
    from sources.spreadsheet_extract import extract_stocklist
    data = "\n".join((
        "Status,Lot,Estate,Suburb,Land,House,Design,Beds,Land Price,Build Price,Package",
        "Available,Lot 82,Aberdeen,Winter Valley,282,142.8,Empley 15,3,$205000,$335220,$540220",
        "Under offer,Lot 83,Aberdeen,Winter Valley,300,150.0,Empley 18,4,$210000,$350000,$560000",
    )).encode("utf-8")
    rows = extract_stocklist(data, "sheet.csv", "Tomorrow Homes")
    assert len(rows) == 2, f"csv gave {len(rows)} listings"
    assert rows[0]["lot_address"] == "Lot 82", rows[0]["lot_address"]
    assert rows[0]["advertised_package_price"] == 540_220
    assert rows[0]["availability_status"] == "Available"
    assert rows[1]["availability_status"] == "Under Offer"
    assert all(r["builder_name"] == "Tomorrow Homes" for r in rows)

    # an HTML sign-in wall where a file was expected must yield nothing, not be parsed
    assert extract_stocklist(b"<!DOCTYPE html><html><body>Sign in to continue", "x") == []


def test_a_wall_of_prices_is_not_a_listing():
    """Real row from E-Agent's commercial page: a transposed sheet whose row is nothing
    but every unit's price. It parsed as a $1,990,000 "listing" with that whole string as
    its address. Rejected only when nothing in the row identifies a property, so the QLD
    dual-occupancy rows — five figures each, but a stock code — are unaffected."""
    from sources.spreadsheet_extract import _listing_from_row

    def kept(text):
        return _listing_from_row(text, [], "", "src", "Builder") is not None

    assert not kept("$1,300,000 $1,250,000 $1,250,000 $1,300,000 $1,990,000 Price "
                    "$650,000 $1,500,000 $1,500,000 $650,000 $1,850,000")
    # five money figures, but identified by a stock code -> a real package
    assert kept("CC-0114 506 Titled 200 5 + 5 + 3 MORETON SINGLE $ 2,380 $ 123,760 8.07% "
                "Download $ 839,000 $ 694,199 $ 1,533,199")
    # ...or by a street address
    assert kept("Woodford 4 Windsor Street Park Rise Jan-27 600 Rectangular $520,000 "
                "Malanda A 4 192.75 4 | 2 | 2 $460,780 $980,780 $5,000 $975,780")
    assert kept("Lot 82 Aberdeen 282 142.8 12.0 Sep-26 Empley 15 3x2x2 $205,000 "
                "$335,220 $540,220 Available")


def test_a_price_broken_out_of_a_listing_is_not_a_listing():
    """Emailed PDFs print a property's figures on their own lines, so "$1,266,900*",
    "Land $714,900" and "Build $552,000" were stored as three separate properties — each
    inheriting the suburb from the heading above, each landing in Coleen's sheet with a
    price in the address column.

    The KEEP cases are why this gate is narrow. A first version tested "has no words" and
    deleted 112 rows, most of them real: an apartment pricelist row is almost entirely
    numeric, so "501 2 1 1 70 27 97 SE $913,000" — unit 501, 2 bed, 1 bath, 1 car, 97 m²,
    south-east — was thrown away. A unit number and a room count are identity even with no
    letters attached."""
    from sources.spreadsheet_extract import _is_price_fragment
    from sources.adaptive_extract import parse_fields
    from sources.feature_extract import parse_listing_features

    def fields(t):
        d = parse_fields(t)
        d.update({k: v for k, v in parse_listing_features(
            t, "", d.get("advertised_package_price")).items() if v is not None})
        return d

    for text in ("$1,266,900*", "Land $714,900", "Build $552,000", "$729,000",
                 "A$744,951", "A$744,951 A$744,951", "Price $1,435,000",
                 "Total $1,120,900 inc GST"):
        assert _is_price_fragment(text, fields(text)), f"kept a fragment: {text!r}"

    for text in ("501 2 1 1 70 27 97 SE $913,000",
                 "104 3 2 2 124 64 188 $2,665,000",
                 "110 1 1 51.1 9.8 60.9 $460,000 S-B4 $2,702",
                 "1404+05 4 4 3 180 18 198 NE $3,200,000",
                 "West 210 Sunrise Deanside Available Paris 22 Elite $412,500 $402,070 "
                 "$814,570",
                 "Lot 82 Aberdeen 282 142.8 12.0 Sep-26 Empley 15 3x2x2 $205,000 $335,220 "
                 "$540,220 Available"):
        assert not _is_price_fragment(text, fields(text)), f"deleted a real listing: {text[:50]!r}"


def test_a_named_column_is_read_instead_of_guessed_from_the_flattened_row():
    """35% of live rows had no bedroom count, 63% no house size — and both were present.

    Every fact is scraped back out of the row once it has been flattened to one string,
    which works for money and fails for bare numbers:

        "AVAILABLE Option 2 Single Kuraby Traditional 10 344.3 164.93 4 2 1 Q4 2026"
                                          frontage ─┘ land ─┘ house ─┘ b b c

    Nothing marks which number is which, so all of them were dropped, and any brief
    stating a minimum then excluded the row. The sheet's own header row names those
    columns, so they are read rather than inferred.
    """
    from sources.spreadsheet_extract import extract_from_csv

    csv_text = (
        "CREATION HOMES NSW STOCK LIST\n"
        "Status,Lot,Option,Storey,House Type,Frontage (m),Land Size (sqm),"
        "House Size (sqm),Bedrooms,Bathrooms,Garage,Registration,Contract,"
        "Land Price,Build Price,Package Price\n"
        "Available,2,N/A,Single,Dualkey - Modern,15.8,502,245.16,6,3.5,2,Q2 2027,"
        "Split,\"$519,000\",\"$732,800\",\"$1,251,800\"\n"
    ).encode()
    rows = extract_from_csv(csv_text, "creation-homes")
    assert len(rows) == 1, rows
    r = rows[0]
    assert r["advertised_package_price"] == 1_251_800, r
    assert r["bedrooms"] == 6 and r["bathrooms"] == 3.5 and r["car_spaces"] == 2, r
    assert r["land_size_sqm"] == 502.0 and r["house_size_sqm"] == 245.16, r
    assert r["frontage_m"] == 15.8, r

    # "Land Price" is money and must never be mistaken for "Land Size".
    assert r["land_size_sqm"] != 519_000

    # A value outside a plausible range means the mapping is wrong for that row, and a
    # wrong bedroom count passes a "minimum 4 bedrooms" filter exactly like an invented
    # one would. It is dropped, not clamped.
    odd = csv_text.replace(b",6,3.5,2,Q2 2027", b",45,3.5,2,Q2 2027")
    assert extract_from_csv(odd, "creation-homes")[0].get("bedrooms") in (None, 0), \
        "45 bedrooms is not a bedroom count"


def run_all():
    tests = [
        ("named columns beat the flattened row",
         test_a_named_column_is_read_instead_of_guessed_from_the_flattened_row),
        ("address column is a label", test_address_column_is_a_label_not_the_whole_row),
        ("column headers are not context", test_column_header_rows_never_become_estate_context),
        ("csv stocklist parsed", test_csv_stocklist_is_parsed),
        ("price fragments are not listings", test_a_price_broken_out_of_a_listing_is_not_a_listing),
        ("wall of prices is not a listing", test_a_wall_of_prices_is_not_a_listing),
        ("xlsx cell hyperlink survives", test_cell_hyperlink_reaches_the_listing),
        ("=HYPERLINK() target recovered", test_hyperlink_formula_target_is_recovered),
        ("rows without links unaffected", test_rows_without_links_are_unaffected),
        ("link kind from anchor or url", test_link_kind_is_taken_from_anchor_or_url),
        ("pdf annots matched by position", test_pdf_annotations_are_matched_to_their_row_by_position),
    ]
    failed = 0
    for name, fn in tests:
        try:
            fn()
            print(f" [PASS] links: {name}")
        except AssertionError as e:
            failed += 1
            print(f" [FAIL] links: {name}: {e}")
    return failed


if __name__ == "__main__":
    import sys
    sys.exit(1 if run_all() else 0)
